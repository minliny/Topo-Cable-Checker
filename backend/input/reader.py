# backend/input/reader.py
# Local input file reader scaffold for Excel/CSV files.
# WARNING: This is a scaffold only. No real business rules applied.
# No database, no AI/LLM, no external API.

import csv
import os
from datetime import datetime
from typing import Optional
from pathlib import Path

from .models import (
    InputFileMetadata,
    RawSheetData,
    RawTabularDataset,
)


class LocalInputReader:
    """Scaffold for reading local Excel/CSV input files.

    WARNING: This is a scaffold only.
    - Does not apply business rules
    - Does not generate issues
    - Does not connect to database
    - Does not use AI/LLM

    Supported formats: .xlsx, .xls, .csv
    Uses openpyxl for Excel files, csv module for CSV.
    """

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

    def read_file(self, file_path: str) -> RawTabularDataset:
        """Read an input file and return raw tabular dataset.

        Args:
            file_path: Path to the input file.

        Returns:
            RawTabularDataset with all sheets/tables.

        Raises:
            FileNotFoundError: If file does not exist.
            ValueError: If file format not supported.
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {file_path}")

        ext = path.suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file format: {ext}. "
                f"Supported: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )

        if ext in (".xlsx", ".xls"):
            return self._read_excel(path)
        else:
            return self._read_csv(path)

    def _read_excel(self, path: Path) -> RawTabularDataset:
        """Read Excel file (scaffold using openpyxl).

        This is a scaffold. Currently reads raw data only.
        No business rules applied.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required to read Excel files. "
                "Install with: pip install openpyxl"
            )

        wb = load_workbook(path, data_only=True, read_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = []

            # Read headers from first row
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell) if cell is not None else "" for cell in row]
                else:
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell for cell in row_data):  # Skip empty rows
                        rows.append(row_data)

            sheets.append(
                RawSheetData(
                    sheet_name=sheet_name,
                    headers=headers,
                    rows=rows,
                    row_count=len(rows),
                    column_count=len(headers),
                )
            )

        wb.close()

        file_size = os.path.getsize(path)
        file_type = path.suffix.lower().lstrip(".")

        metadata = InputFileMetadata(
            file_path=str(path.absolute()),
            file_name=path.name,
            file_size=file_size,
            file_type=file_type,
            sheet_count=len(sheets),
        )

        return RawTabularDataset(
            metadata=metadata,
            sheets=sheets,
            source_file=str(path.absolute()),
        )

    def _read_csv(self, path: Path) -> RawTabularDataset:
        """Read CSV file."""
        rows = []
        headers = []

        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader):
                if row_idx == 0:
                    headers = row
                else:
                    if any(cell for cell in row):
                        rows.append(row)

        sheets = [
            RawSheetData(
                sheet_name="data",
                headers=headers,
                rows=rows,
                row_count=len(rows),
                column_count=len(headers),
            )
        ]

        file_size = os.path.getsize(path)

        metadata = InputFileMetadata(
            file_path=str(path.absolute()),
            file_name=path.name,
            file_size=file_size,
            file_type="csv",
            sheet_count=1,
        )

        return RawTabularDataset(
            metadata=metadata,
            sheets=sheets,
            source_file=str(path.absolute()),
        )

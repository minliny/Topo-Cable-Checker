import openpyxl

wb = openpyxl.Workbook()

# Devices Sheet
ws_devices = wb.active
ws_devices.title = "Devices"
ws_devices.append(["Device Name", "Device Type", "Status"])
ws_devices.append(["SW-01", "Switch", "active"])
ws_devices.append(["SW-02", "Switch", "offline"]) # field_equals violation
ws_devices.append(["RTR-ABC", "Router", "active"]) # regex violation
ws_devices.append(["", "Switch", "active"]) # missing_value violation

# Ports Sheet
ws_ports = wb.create_sheet("Ports")
ws_ports.append(["Device Name", "Port Name", "Port Status"])
ws_ports.append(["SW-01", "Eth1/1", "up"])
ws_ports.append(["SW-01", "Eth1/2", "down"])

# Links Sheet
ws_links = wb.create_sheet("Links")
ws_links.append(["Source Device", "Source Port", "Dest Device", "Dest Port"])
ws_links.append(["SW-01", "Eth1/1", "SW-02", "Eth1/1"])

wb.save("test_network_data.xlsx")
print("test_network_data.xlsx generated")

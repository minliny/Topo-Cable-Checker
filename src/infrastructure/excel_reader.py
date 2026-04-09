import openpyxl
from typing import Dict, List, Any
import os

class ExcelReader:
    def read(self, file_path: str) -> Dict[str, Any]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
            
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = {
            "sheets": wb.sheetnames,
            "header_mapping": {},
            "row_data": {}
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
                
            headers = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(rows[0])]
            result["header_mapping"][sheet_name] = headers
            
            sheet_data = []
            for row in rows[1:]:
                # skip completely empty rows
                if all(cell is None for cell in row):
                    continue
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell
                sheet_data.append(row_dict)
                
            result["row_data"][sheet_name] = sheet_data
            
        return result

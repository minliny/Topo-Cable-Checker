import openpyxl

wb = openpyxl.Workbook()

# Devices Sheet
ws_devices = wb.active
ws_devices.title = "Devices"
ws_devices.append(["Device Name", "Device Type", "Status"])
ws_devices.append(["SW-01", "Switch", "active"])
ws_devices.append(["SW-02", "Switch", "active"]) # RESOLVED: fixed status
ws_devices.append(["RTR-ABC", "Router", "offline"]) # PERSISTENT & CHANGED: still regex fail, but status changed to offline
# Removed the empty device, so it is RESOLVED
ws_devices.append(["SW-03", "Switch", "offline"]) # NEW: status is offline

# Ports Sheet
ws_ports = wb.create_sheet("Ports")
ws_ports.append(["Device Name", "Port Name", "Port Status"])
ws_ports.append(["SW-01", "Eth1/1", "up"])

# Links Sheet
ws_links = wb.create_sheet("Links")
ws_links.append(["Source Device", "Source Port", "Dest Device", "Dest Port"])
ws_links.append(["SW-01", "Eth1/1", "SW-02", "Eth1/1"])

wb.save("test_network_data_v2.xlsx")
print("test_network_data_v2.xlsx generated")

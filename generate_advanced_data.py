import openpyxl

wb = openpyxl.Workbook()

# Devices Sheet
ws_devices = wb.active
ws_devices.title = "Devices"
ws_devices.append(["Device Name", "Device Type", "Status"])
ws_devices.append(["SW-01", "Switch", "active"])
ws_devices.append(["SW-02", "Switch", "active"]) 
ws_devices.append(["SW-03", "Switch", "offline"]) # Group consistency failure (type Switch, should be active)
ws_devices.append(["ISOLATED-01", "Router", "active"]) # Isolated device
# Missing peer target device RTR-PEER

# Ports Sheet
ws_ports = wb.create_sheet("Ports")
ws_ports.append(["Device Name", "Port Name", "Port Status"])
ws_ports.append(["SW-01", "Eth1/1", "up"])

# Links Sheet
ws_links = wb.create_sheet("Links")
ws_links.append(["Source Device", "Source Port", "Dest Device", "Dest Port"])
ws_links.append(["SW-01", "Eth1/1", "SW-02", "Eth1/1"])
ws_links.append(["SW-01", "Eth1/1", "SW-02", "Eth1/1"]) # Duplicate link
ws_links.append(["SW-01", "Eth1/2", "RTR-PEER", "Eth1/1"]) # Missing peer (RTR-PEER)
ws_links.append(["SW-02", "Eth1/2", "SW-02", "Eth1/3"]) # Self loop

wb.save("test_advanced_data.xlsx")
print("test_advanced_data.xlsx generated")
